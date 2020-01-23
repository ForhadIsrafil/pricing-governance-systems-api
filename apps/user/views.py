from rest_framework.views import APIView
from rest_framework import status
from .models import User, UserInfo
from .serializers import UserSerializer, AuthTokenSerializer, UserInfoSerializer
from const import slicer
from django.http import Http404
from django.contrib.auth import authenticate
from rest_framework.authtoken.models import Token
from rest_framework.compat import coreapi, coreschema
from rest_framework.response import Response
from rest_framework.schemas import ManualSchema
from rest_framework import parsers, renderers, status
import json
from django.db import transaction, DatabaseError
import os
import hashlib
import tempfile
import shutil
from django.conf import settings
import uuid
import pandas as pd
import xlrd
from openpyxl import load_workbook
from data.models import PocketMarginData, YearInfo
import permissions
from threading import Thread
from time import sleep
from .user_query_manager import *


TOTAL_DATA_TO_READ = 0
CURRENT_READ_DATA = 2   # as our row started at this point
PROGRESS = 0
RESET = 0
PROCESS_FAIL = 0


class ObtainAuthToken(APIView):
    throttle_classes = ()
    permission_classes = ()
    parser_classes = (parsers.FormParser, parsers.MultiPartParser, parsers.JSONParser,)
    renderer_classes = (renderers.JSONRenderer,)
    serializer_class = AuthTokenSerializer
    if coreapi is not None and coreschema is not None:
        schema = ManualSchema(
            fields=[
                coreapi.Field(
                    name="username",
                    required=True,
                    location='form',
                    schema=coreschema.String(
                        title="Username",
                        description="Valid username for authentication",
                    ),
                ),
                coreapi.Field(
                    name="password",
                    required=True,
                    location='form',
                    schema=coreschema.String(
                        title="Password",
                        description="Valid password for authentication",
                    ),
                ),
            ],
            encoding="application/json",
        )

    def post(self, request, *args, **kwargs):

        req_data = get_login_data_formate(request.data)
        serializer = self.serializer_class(data=req_data,
                                           context={'request': request})
        if serializer.is_valid():

            # making deactive user active
            response = set_user_active(serializer.data)

            if response == False:
                return Response(status=status.HTTP_409_CONFLICT)

            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class Signout(APIView):
    def post(self, request, format=None):
        request.user.auth_token.delete()
        return Response(status=status.HTTP_200_OK)


class NewUserView(APIView):
    permission_classes = (permissions.IsAuthenticated, permissions.IsAdminUser,)

    def post(self, request):

        req_data = get_user_data(request.data)

        serializer = UserInfoSerializer(data= req_data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UploadFileView(APIView):
    permission_classes = (permissions.IsAuthenticated, permissions.IsAdminUser,)

    def post(self, request, format=None):

        year = request.POST.get('year')

        file = request.FILES.get('file')
        if None in [file, year]:
            return Response(status=status.HTTP_400_BAD_REQUEST)


        trimed_year = year.strip()
        if trimed_year == '':
            return Response(status=status.HTTP_400_BAD_REQUEST)

        hash = str(uuid.uuid4())

        # Process the uploaded model
        _, ext = os.path.splitext(file.name)
        type = ext[1:].lower() if len(ext) > 0 else None

        with tempfile.NamedTemporaryFile(delete=False) as fp:
            tmppath = fp.name
            for chunk in file.chunks():
                fp.write(chunk)

            # Save the model in the static path
            path = settings.FILE_ROOT + '/' + hash + '.' + type
            shutil.copyfile(tmppath, path)

        file_name = hash + '.' + type
        save_data_thread = Thread(target=save_pocket_margin_data, args=(year, file_name,))

        # starting thread 1
        save_data_thread.start()

        return Response(status=status.HTTP_201_CREATED)


class GetProgress(APIView):
    def get(self, request, format=None):
        global CURRENT_READ_DATA, TOTAL_DATA_TO_READ, PROGRESS, RESET, PROCESS_FAIL

        if PROCESS_FAIL== 1:
            PROCESS_FAIL = 0
            return Response(status=status.HTTP_400_BAD_REQUEST)

        if RESET == 1:
            TOTAL_DATA_TO_READ = 0
            CURRENT_READ_DATA = 2
            RESET = 0

        if PROGRESS == 100:
            PROGRESS = 0 # reset progress to begining
            return Response(100, status=status.HTTP_200_OK)

        try:
            progress = (CURRENT_READ_DATA / TOTAL_DATA_TO_READ) * 100.0
        except:
            return Response(0, status=status.HTTP_200_OK)
        return Response(progress, status=status.HTTP_200_OK)


def save_pocket_margin_data(year, file_name):
    global CURRENT_READ_DATA, RESET, PROGRESS, PROCESS_FAIL
    try:

        file = settings.FILE_ROOT + '/' + file_name
        wb = load_workbook(file)

        names = wb.get_sheet_names()
        sheet = wb.get_sheet_by_name(names[0])
        max_row = sheet.max_row

        # setting global variable
        global TOTAL_DATA_TO_READ
        TOTAL_DATA_TO_READ = max_row - 1

        max_column = sheet.max_column

        year_info = YearInfo(year=year, file_name=file_name)   # default = 2 (Pending)
        year_info.save()

        for row in range(3, max_row):
            table = PocketMarginData(year_info=year_info)

            # setting global variable

            CURRENT_READ_DATA += 1

            if CURRENT_READ_DATA == TOTAL_DATA_TO_READ:     # target completed
                PROGRESS = 100
                RESET = 1
                TOTAL_DATA_TO_READ = 0
                CURRENT_READ_DATA = 2

            # checking empty row
            if sheet.cell(column=1, row=row).value == None:
                continue

            for column in range(1, 38):
                data = sheet.cell(column=column, row=row).value
                if column == 1:
                    table.profit_center_code = data
                elif column == 2:
                    table.profit_center_name = data
                elif column == 3:
                    table.product_segmentation = data
                elif column == 4:
                    table.channel_partnar_ind = data
                elif column == 5:
                    table.business_segment_group = data
                elif column == 6:
                    table.customer_group_code = data
                elif column == 7:
                    table.customer_group_name = data
                elif column == 8:
                    table.material_number_code = data
                elif column == 9:
                    table.material_name = data
                elif column == 10:
                    table.freight_type_code = data
                elif column == 11:
                    table.freight_types = data
                elif column == 12:
                    table.best_fit_acc_man = data
                elif column == 13:
                    table.manf_plant = data
                elif column == 14:
                    table.sold_to_region = data
                elif column == 15:
                    table.business_segment = data
                elif column == 16:
                    table.product_family = data
                elif column == 17:
                    table.sales_volume_mt = data
                elif column == 18:
                    table.gross_sale_usd = data
                elif column == 19:
                    table.invoice_price = data
                elif column == 20:
                    table.freight_costs = data
                elif column == 21:
                    table.freight_revenue = data
                elif column == 22:
                    table.other_discounts_and_rebates = data
                elif column == 23:
                    table.pocket_price = data
                elif column == 24:
                    table.cogs = data
                elif column == 25:
                    table.pocket_margin = data
                elif column == 26:
                    table.pocket_margin_percentage = data * 100
                elif column == 27:
                    table.volume_bands = data
                elif column == 28:
                    table.floor_pocket_margin_corresponding_band = data
                elif column == 29:
                    table.target_pocket_margin_corresponding_band = data
                elif column == 30:
                    table.lower_than_floor_flag = data
                elif column == 31:
                    table.lower_than_target_flag = data
                elif column == 32:
                    table.change_in_invoice_price_per_mt_if_using_floor = data
                elif column == 33:
                    table.change_in_invoice_price_per_mt_if_using_target = data
                elif column == 34:
                    table.opportunity_to_floor = data
                elif column == 35:
                    table.opportunity_to_target = data
                elif column == 36:
                    table.percentage_change_in_invoice_price_or_mt_if_using_floor_margin = data
                elif column == 37:
                    table.percentage_change_in_invoice_price_or_mt_if_using_target_margin = data

            table.save()

        year_info.status = 0   # making this draft  (temporary)
        year_info.save()


        return True

    except:
        TOTAL_DATA_TO_READ = 0
        CURRENT_READ_DATA = 2
        PROGRESS = 0
        RESET = 0
        PROCESS_FAIL = 1

        return False
